import io
import uuid
from pathlib import Path
from typing import Dict, Any, List, Optional
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.core.config import settings
from app.schemas.profiling import DatasetProfileResponse
from app.schemas.anomaly import AnomalyResultsResponse


class ExcelExporter:
    # Color palette
    NAVY_HEADER = "1E293B"
    BLUE_ACCENT = "2563EB"
    LIGHT_GRAY = "F8FAFC"
    BORDER_COLOR = "CBD5E1"
    ALERT_RED = "FEE2E2"
    ALERT_YELLOW = "FEF3C7"
    ALERT_GREEN = "DCFCE7"

    @classmethod
    def generate_comprehensive_workbook(
        cls,
        df_cleaned: pd.DataFrame,
        profile: DatasetProfileResponse,
        anomalies_res: Optional[AnomalyResultsResponse] = None,
        dataset_name: str = "Dataset"
    ) -> Path:
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Style objects
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=cls.NAVY_HEADER, end_color=cls.NAVY_HEADER, fill_type="solid")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
        bold_font = Font(name="Calibri", size=10, bold=True)
        regular_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style='thin', color=cls.BORDER_COLOR),
            right=Side(style='thin', color=cls.BORDER_COLOR),
            top=Side(style='thin', color=cls.BORDER_COLOR),
            bottom=Side(style='thin', color=cls.BORDER_COLOR)
        )

        # -------------------------------------------------------------
        # Sheet 1: Cleaned Data
        # -------------------------------------------------------------
        ws_data = wb.create_sheet(title="Cleaned Data")
        headers = list(df_cleaned.columns)
        ws_data.append(headers)

        for col_num in range(1, len(headers) + 1):
            cell = ws_data.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, row in df_cleaned.iterrows():
            row_vals = [None if pd.isna(v) else v for v in row.values]
            ws_data.append(row_vals)

        # Auto column width
        for col in ws_data.columns:
            max_len = max(len(str(cell.value or '')) for cell in col[:100])
            col_letter = get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = max(12, min(max_len + 4, 45))

        # -------------------------------------------------------------
        # Sheet 2: Executive Summary & Health Score
        # -------------------------------------------------------------
        ws_summary = wb.create_sheet(title="Executive Summary")
        ws_summary.append([f"Quantura | Executive Audit Summary for {dataset_name}"])
        ws_summary.cell(row=1, column=1).font = title_font
        ws_summary.append([f"Health Score: {profile.health.overall_score}/100 | Total Rows: {profile.total_rows:,} | Total Columns: {profile.total_columns}"])
        ws_summary.cell(row=2, column=1).font = sub_font
        ws_summary.append([])

        ws_summary.append(["Health Category", "Score (0-100)", "Assessment"])
        for col_num in range(1, 4):
            c = ws_summary.cell(row=4, column=col_num)
            c.font = header_font
            c.fill = header_fill

        categories = [
            ("Completeness", profile.health.completeness_score, "High completeness" if profile.health.completeness_score > 90 else "Review missing fields"),
            ("Uniqueness", profile.health.uniqueness_score, "No duplicate inflation" if profile.health.uniqueness_score > 95 else "Duplicate records detected"),
            ("Consistency", profile.health.consistency_score, "Standardized format"),
            ("Validity", profile.health.validity_score, "Meets schema constraints"),
            ("Anomaly Risk", profile.health.anomaly_risk_score, "Low anomaly exposure" if profile.health.anomaly_risk_score > 80 else "Review flagged outlier records")
        ]

        for cat, score, assessment in categories:
            ws_summary.append([cat, f"{score}%", assessment])

        # -------------------------------------------------------------
        # Sheet 3: Detected Anomalies
        # -------------------------------------------------------------
        ws_anom = wb.create_sheet(title="Anomalies")
        anom_headers = ["Row #", "Column", "Actual Value", "Expected Range", "Deviation", "Score", "Severity", "Reason"]
        ws_anom.append(anom_headers)
        for col_num in range(1, len(anom_headers) + 1):
            c = ws_anom.cell(row=1, column=col_num)
            c.font = header_font
            c.fill = header_fill

        if anomalies_res and anomalies_res.anomalies:
            for a in anomalies_res.anomalies:
                ws_anom.append([
                    a.row_index,
                    a.column_name,
                    str(a.actual_value),
                    a.expected_range,
                    f"{a.deviation_pct:+}%" if a.deviation_pct is not None else "N/A",
                    a.anomaly_score,
                    a.severity.upper(),
                    a.reason
                ])
        else:
            ws_anom.append(["-", "No high-risk anomalies detected", "-", "-", "-", "-", "-", "-"])

        # -------------------------------------------------------------
        # Sheet 4: Column Statistics
        # -------------------------------------------------------------
        ws_stats = wb.create_sheet(title="Column Statistics")
        stats_headers = ["Column Name", "Detected Type", "Confidence", "Missing Count", "Missing %", "Unique Values", "Quality Score"]
        ws_stats.append(stats_headers)
        for col_num in range(1, len(stats_headers) + 1):
            c = ws_stats.cell(row=1, column=col_num)
            c.font = header_font
            c.fill = header_fill

        for col in profile.columns:
            ws_stats.append([
                col.name,
                col.detected_type.title(),
                f"{int(col.confidence * 100)}%",
                col.null_count,
                f"{col.null_percentage}%",
                col.unique_count,
                f"{col.quality_score}/100"
            ])

        # Save workbook to exports storage
        export_filename = f"Quantura_{dataset_name.replace(' ', '_')}_{uuid.uuid4().hex[:8]}.xlsx"
        export_path = settings.EXPORTS_PATH / export_filename
        wb.save(export_path)

        return export_path
