import os
import uuid
import re
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
import pandas as pd
import numpy as np
import openpyxl
from app.core.config import settings
from app.schemas.dataset import SheetInfo, FileInspectResponse


class FileService:
    @staticmethod
    def format_file_size(size_bytes: int) -> str:
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024.0:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024.0
        return f"{size_bytes:.1f} TB"

    @staticmethod
    def save_upload_file(filename: str, content: bytes) -> Tuple[str, str, int, str]:
        ext = Path(filename).suffix.lower()
        if ext not in [".csv", ".xlsx", ".xls"]:
            raise ValueError(f"Unsupported file extension: {ext}. Supported formats are .csv, .xlsx, .xls")
        
        size_bytes = len(content)
        max_bytes = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024
        if size_bytes > max_bytes:
            raise ValueError(f"File size exceeds maximum allowed limit of {settings.MAX_UPLOAD_SIZE_MB}MB")

        stored_filename = f"{uuid.uuid4()}{ext}"
        target_path = settings.DATASETS_PATH / stored_filename

        with open(target_path, "wb") as f:
            f.write(content)

        file_type = ext.replace(".", "")
        return stored_filename, file_type, size_bytes, str(target_path)

    @staticmethod
    def inspect_file(file_path: Path, original_filename: str) -> FileInspectResponse:
        ext = file_path.suffix.lower()
        file_size = file_path.stat().st_size
        formatted_size = FileService.format_file_size(file_size)

        if ext == ".csv":
            # Try reading sample
            for encoding in ["utf-8", "utf-8-sig", "latin1", "cp1252"]:
                try:
                    df = pd.read_csv(file_path, nrows=50, encoding=encoding)
                    total_rows = sum(1 for _ in open(file_path, "r", encoding=encoding, errors="ignore")) - 1
                    total_rows = max(0, total_rows)
                    
                    # Clean columns
                    cols = [str(c) if str(c).strip() else f"Unnamed_{i}" for i, c in enumerate(df.columns)]
                    df.columns = cols
                    preview_records = df.head(10).replace({np.nan: None}).to_dict(orient="records")

                    sheet_info = SheetInfo(
                        name="Sheet1",
                        row_count=total_rows,
                        column_count=len(cols),
                        columns=cols,
                        preview_rows=preview_records
                    )
                    return FileInspectResponse(
                        filename=original_filename,
                        file_type="csv",
                        file_size_bytes=file_size,
                        file_size_formatted=formatted_size,
                        sheet_count=1,
                        sheet_names=["Sheet1"],
                        sheets=[sheet_info]
                    )
                except Exception:
                    continue
            raise ValueError("Could not read CSV file. Please ensure valid encoding (UTF-8, Latin-1).")

        elif ext in [".xlsx", ".xls"]:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            sheets: List[SheetInfo] = []

            for name in sheet_names:
                ws = wb[name]
                rows_iter = ws.iter_rows(values_only=True)
                header_row = next(rows_iter, None)
                if not header_row:
                    continue
                
                cols = [str(c) if c is not None and str(c).strip() else f"Column_{i+1}" for i, c in enumerate(header_row)]
                preview_list = []
                row_count = 0
                for row_idx, r in enumerate(rows_iter):
                    row_count += 1
                    if row_idx < 10:
                        row_dict = {}
                        for c_idx, val in enumerate(r):
                            if c_idx < len(cols):
                                row_dict[cols[c_idx]] = val
                        preview_list.append(row_dict)
                
                # Estimate total row count
                total_est = ws.max_row - 1 if ws.max_row and ws.max_row > 1 else row_count
                sheets.append(SheetInfo(
                    name=name,
                    row_count=total_est,
                    column_count=len(cols),
                    columns=cols,
                    preview_rows=preview_list
                ))
            wb.close()
            
            return FileInspectResponse(
                filename=original_filename,
                file_type=ext.replace(".", ""),
                file_size_bytes=file_size,
                file_size_formatted=formatted_size,
                sheet_count=len(sheet_names),
                sheet_names=sheet_names,
                sheets=sheets
            )
        else:
            raise ValueError(f"Unsupported file format {ext}")

    @staticmethod
    def load_dataframe(file_path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
        ext = file_path.suffix.lower()
        if ext == ".csv":
            for enc in ["utf-8", "utf-8-sig", "latin1", "cp1252"]:
                try:
                    df = pd.read_csv(file_path, encoding=enc)
                    return df
                except Exception:
                    continue
            raise ValueError("Failed to load CSV file with supported encodings")
        elif ext in [".xlsx", ".xls"]:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            return df
        else:
            raise ValueError(f"Unsupported file type {ext}")

    @staticmethod
    def save_dataframe_to_version(df: pd.DataFrame, ext: str = ".csv") -> Tuple[str, Path, int]:
        new_filename = f"{uuid.uuid4()}{ext}"
        new_path = settings.DATASETS_PATH / new_filename
        if ext == ".csv":
            df.to_csv(new_path, index=False)
        else:
            df.to_excel(new_path, index=False)
        size_bytes = new_path.stat().st_size
        return new_filename, new_path, size_bytes
